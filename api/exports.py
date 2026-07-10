"""Excel 导出 + 个人统计详情 API"""
from datetime import date, datetime
from io import BytesIO
from flask import request, jsonify, send_file
from api import api_bp
from database import db
from models import User, Project, TimeEntry, TimeEntryDetail, LeaveRecord, Department
from utils.holiday import get_workdays_count, get_working_hours
from config import STANDARD_HOURS_PER_DAY
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ==================== 工具：样式 ====================
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_cell(ws, row, col, align=CENTER_ALIGN):
    cell = ws.cell(row=row, column=col)
    cell.alignment = align
    cell.border = THIN_BORDER
    return cell


# ==================== 日表导出 ====================
@api_bp.route("/exports/daily", methods=["GET"])
def export_daily():
    date_str = request.args.get("date") or date.today().isoformat()
    try:
        report_date = date.fromisoformat(date_str)
    except ValueError:
        return jsonify({"code": 400, "msg": "日期格式错误"})

    users = User.query.filter_by(is_active=True).order_by(User.department_id, User.name).all()
    projects = Project.query.filter_by(is_active=True).order_by(Project.code).all()
    leaves = LeaveRecord.query.filter_by(leave_date=report_date).all()
    leave_user_ids = set(l.user_id for l in leaves)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"日报_{date_str}"

    # 标题行
    headers = ["人员", "中心", "项目", "任务", "进度说明", "工时(h)", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for u in users:
        has_entry = False
        for p in projects:
            entry = TimeEntry.query.filter_by(
                user_id=u.id, project_id=p.id, record_date=report_date
            ).first()
            if entry:
                has_entry = True
                for d in entry.details:
                    _style_cell(ws, row, 1).value = u.name
                    _style_cell(ws, row, 2).value = u.department.name if u.department else ""
                    _style_cell(ws, row, 3).value = p.code
                    _style_cell(ws, row, 4).value = d.task.task_name if d.task else ""
                    _style_cell(ws, row, 5).value = d.progress
                    _style_cell(ws, row, 6).value = d.hours
                    _style_cell(ws, row, 7).value = entry.remark or ""
                    row += 1

        if not has_entry and u.id in leave_user_ids:
            _style_cell(ws, row, 1).value = u.name
            _style_cell(ws, row, 2).value = u.department.name if u.department else ""
            _style_cell(ws, row, 3).value = "请假"
            _style_cell(ws, row, 4).value = "-"
            _style_cell(ws, row, 5).value = 0
            _style_cell(ws, row, 6).value = 0
            _style_cell(ws, row, 7).value = "请假"
            row += 1

    # 列宽
    widths = [12, 14, 12, 18, 10, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"日报_{date_str}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 月表导出 ====================
@api_bp.route("/exports/monthly", methods=["GET"])
def export_monthly():
    year = request.args.get("year", type=int) or date.today().year
    month = request.args.get("month", type=int) or date.today().month
    user_id = request.args.get("user_id", type=int)

    query = TimeEntry.query.filter(
        db.extract("year", TimeEntry.record_date) == year,
        db.extract("month", TimeEntry.record_date) == month,
    )
    if user_id:
        query = query.filter_by(user_id=user_id)
    entries = query.order_by(TimeEntry.user_id, TimeEntry.record_date).all()

    # 汇总
    summary = {}
    for e in entries:
        key = (e.user_id, e.project_id)
        total = sum(d.hours for d in e.details)
        summary[key] = summary.get(key, 0) + total

    users_data = User.query.filter_by(is_active=True).all() if not user_id else [User.query.get(user_id)]
    users_data = [u for u in users_data if u]
    work_days = get_workdays_count(year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"月报_{year}_{month}"

    headers = ["人员", "中心", "项目", "总工时(h)", "工作日(天)", "请假(天)", "出勤(天)", "标准工时(h)", "达成率(%)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for u in users_data:
        leave_days = LeaveRecord.query.filter(
            LeaveRecord.user_id == u.id,
            db.extract("year", LeaveRecord.leave_date) == year,
            db.extract("month", LeaveRecord.leave_date) == month,
        ).count()
        user_work_days = work_days - leave_days
        user_standard_hours = user_work_days * STANDARD_HOURS_PER_DAY

        total_hours = 0
        for (uid, pid), hrs in sorted(summary.items()):
            if uid == u.id:
                p = Project.query.get(pid)
                _style_cell(ws, row, 1).value = u.name
                _style_cell(ws, row, 2).value = u.department.name if u.department else ""
                _style_cell(ws, row, 3).value = p.code if p else ""
                _style_cell(ws, row, 4).value = round(hrs, 1)
                _style_cell(ws, row, 5).value = work_days
                _style_cell(ws, row, 6).value = leave_days
                _style_cell(ws, row, 7).value = user_work_days
                _style_cell(ws, row, 8).value = user_standard_hours
                pct = round((hrs / user_standard_hours * 100), 1) if user_standard_hours > 0 else 0
                _style_cell(ws, row, 9).value = pct
                row += 1
                total_hours += hrs

        # 如果该用户本月没任何工时记录，也显示一行
        if total_hours == 0:
            _style_cell(ws, row, 1).value = u.name
            _style_cell(ws, row, 2).value = u.department.name if u.department else ""
            _style_cell(ws, row, 3).value = "-"
            _style_cell(ws, row, 4).value = 0
            _style_cell(ws, row, 5).value = work_days
            _style_cell(ws, row, 6).value = leave_days
            _style_cell(ws, row, 7).value = user_work_days
            _style_cell(ws, row, 8).value = user_standard_hours
            _style_cell(ws, row, 9).value = 0
            row += 1

    widths = [12, 14, 14, 12, 12, 10, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"月报_{year}_{month}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== 个人统计详情 ====================
@api_bp.route("/reports/person-detail", methods=["GET"])
def person_detail():
    """管理员查看某个人员的详细统计"""
    user_id = request.args.get("user_id", type=int)
    year = request.args.get("year", type=int) or date.today().year
    month = request.args.get("month", type=int) or date.today().month

    user = User.query.get_or_404(user_id)

    # 当月每天工时明细
    entries = TimeEntry.query.filter(
        TimeEntry.user_id == user_id,
        db.extract("year", TimeEntry.record_date) == year,
        db.extract("month", TimeEntry.record_date) == month,
    ).order_by(TimeEntry.record_date, TimeEntry.project_id).all()

    # 按天+项目汇总
    daily = {}
    daily_total = {}
    for e in entries:
        day_key = e.record_date.isoformat()
        if day_key not in daily:
            daily[day_key] = []
            daily_total[day_key] = 0
        proj_hours = sum(d.hours for d in e.details)
        daily[day_key].append({
            "project_code": e.project.code if e.project else "",
            "project_name": e.project.name if e.project else "",
            "total_hours": round(proj_hours, 1),
            "details": [{
                "task_name": d.task.task_name if d.task else "",
                "progress": d.progress,
                "hours": d.hours,
            } for d in e.details],
        })
        daily_total[day_key] += proj_hours

    # 项目汇总
    proj_summary = {}
    for e in entries:
        pid = e.project_id
        hrs = sum(d.hours for d in e.details)
        proj_summary[pid] = proj_summary.get(pid, 0) + hrs

    project_list = []
    for pid, total_hours in proj_summary.items():
        p = Project.query.get(pid)
        project_list.append({
            "project_id": pid,
            "project_code": p.code if p else "",
            "project_name": p.name if p else "",
            "total_hours": round(total_hours, 1),
        })

    # 请假记录
    leaves = LeaveRecord.query.filter(
        LeaveRecord.user_id == user_id,
        db.extract("year", LeaveRecord.leave_date) == year,
        db.extract("month", LeaveRecord.leave_date) == month,
    ).order_by(LeaveRecord.leave_date).all()

    # 统计汇总
    work_days = get_workdays_count(year, month)
    leave_days = len(leaves)
    actual_work_days = work_days - leave_days
    standard_hours = actual_work_days * STANDARD_HOURS_PER_DAY
    total_hours = sum(daily_total.values())
    percentage = round((total_hours / standard_hours * 100), 1) if standard_hours > 0 else 0

    # 日列表（含周末/节假日标记）
    from models import WorkingDay
    cal_days = WorkingDay.query.filter(
        db.extract("year", WorkingDay.date) == year,
        db.extract("month", WorkingDay.date) == month,
    ).order_by(WorkingDay.date).all()

    calendar = []
    for cd in cal_days:
        day_str = cd.date.isoformat()
        is_leave = any(l.leave_date.isoformat() == day_str for l in leaves)
        calendar.append({
            "date": day_str,
            "is_workday": cd.is_workday,
            "holiday_name": cd.holiday_name,
            "is_leave": is_leave,
            "has_entry": day_str in daily,
            "total_hours": daily_total.get(day_str, 0),
            "projects": daily.get(day_str, []),
        })

    return jsonify({
        "code": 0,
        "data": {
            "user": {
                "id": user.id,
                "name": user.name,
                "department_name": user.department.name if user.department else "",
                "role": user.role,
            },
            "summary": {
                "year": year,
                "month": month,
                "work_days": work_days,
                "leave_days": leave_days,
                "actual_work_days": actual_work_days,
                "standard_hours": standard_hours,
                "total_hours": round(total_hours, 1),
                "percentage": percentage,
            },
            "projects": sorted(project_list, key=lambda x: x["project_code"]),
            "calendar": calendar,
            "leaves": [{
                "id": l.id,
                "leave_date": l.leave_date.isoformat(),
                "leave_type": l.leave_type,
                "remark": l.remark,
            } for l in leaves],
        }
    })


# ==================== 个人月报导出 ====================
@api_bp.route("/exports/person", methods=["GET"])
def export_person():
    """导出某个人员的当月详细工时 Excel"""
    user_id = request.args.get("user_id", type=int)
    year = request.args.get("year", type=int) or date.today().year
    month = request.args.get("month", type=int) or date.today().month

    user = User.query.get_or_404(user_id)

    entries = TimeEntry.query.filter(
        TimeEntry.user_id == user_id,
        db.extract("year", TimeEntry.record_date) == year,
        db.extract("month", TimeEntry.record_date) == month,
    ).order_by(TimeEntry.record_date, TimeEntry.project_id).all()

    # 汇总
    proj_totals = {}
    for e in entries:
        pid = e.project_id
        hrs = sum(d.hours for d in e.details)
        proj_totals[pid] = proj_totals.get(pid, 0) + hrs

    leaves = LeaveRecord.query.filter(
        LeaveRecord.user_id == user_id,
        db.extract("year", LeaveRecord.leave_date) == year,
        db.extract("month", LeaveRecord.leave_date) == month,
    ).all()
    work_days = get_workdays_count(year, month)
    leave_days = len(leaves)

    wb = openpyxl.Workbook()
    # Sheet 1: 每日明细
    ws = wb.active
    ws.title = "每日明细"
    headers = ["日期", "项目", "任务", "进度说明", "工时(h)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for e in entries:
        for d in e.details:
            _style_cell(ws, row, 1).value = e.record_date.isoformat()
            _style_cell(ws, row, 2).value = e.project.code if e.project else ""
            _style_cell(ws, row, 3).value = d.task.task_name if d.task else ""
            _style_cell(ws, row, 4).value = d.progress
            _style_cell(ws, row, 5).value = d.hours
            row += 1

    for l in leaves:
        _style_cell(ws, row, 1).value = l.leave_date.isoformat()
        _style_cell(ws, row, 2).value = "请假"
        _style_cell(ws, row, 3).value = l.leave_type
        _style_cell(ws, row, 4).value = 0
        _style_cell(ws, row, 5).value = 0
        row += 1

    for i, w in enumerate([14, 12, 18, 10, 10], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Sheet 2: 汇总
    ws2 = wb.create_sheet("月度汇总")
    headers2 = ["项目", "总工时(h)", "工作日(天)", "请假(天)", "出勤(天)", "标准工时(h)", "达成率(%)"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i, value=h)
    _style_header(ws2, 1, len(headers2))

    row2 = 2
    for pid, hrs in sorted(proj_totals.items()):
        p = Project.query.get(pid)
        actual_days = work_days - leave_days
        std_hours = actual_days * STANDARD_HOURS_PER_DAY
        pct = round((hrs / std_hours * 100), 1) if std_hours > 0 else 0
        _style_cell(ws2, row2, 1).value = p.code if p else ""
        _style_cell(ws2, row2, 2).value = round(hrs, 1)
        _style_cell(ws2, row2, 3).value = work_days
        _style_cell(ws2, row2, 4).value = leave_days
        _style_cell(ws2, row2, 5).value = work_days - leave_days
        _style_cell(ws2, row2, 6).value = std_hours
        _style_cell(ws2, row2, 7).value = pct
        row2 += 1

    for i, w in enumerate([14, 12, 12, 10, 10, 12, 12], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"{user.name}_{year}_{month}_工时明细.xlsx"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
